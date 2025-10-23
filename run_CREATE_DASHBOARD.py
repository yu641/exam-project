import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
import logging

# --- 로깅 설정 ---
log_dash = logging.getLogger("dashboard_creator")
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


# --- (필수) 설정 ---

# 1. 분석할 채점 완료 파일
GRADED_FILE_PATH = r".\output\EX20251021T071942-S001-EXAM_graded.xlsx"

# 2. 원본 DB 파일
DB_PATH = r".\data\db_with_irt_from_distractors.xlsx"

# 3. 생성될 대시보드 파일 경로
DASHBOARD_OUTPUT_PATH = r".\output\DASHBOARD_S001_20251023_1.xlsx"

# 4. 독서/문학 대분류 매핑
LIT_SUBJECTS = ["현대시", "현대시*", "현대소설", "고전시가", "고전시가*", "고전소설"]
READ_SUBJECTS = ["인문", "주제통합", "예술", "과학", "기술", "과학기술", "과학·기술", "사회"]

# --- (필수) 취약점 분석 기준 ---
WEAK_PASSAGE_THRESHOLD = 70.0 # 이 정답률 미만일 시 '취약 지문'
WEAK_PROBLEM_THRESHOLD = 60.0 # 이 정답률 미만일 시 '취약 문제'

# --- 스타일 정의 ---
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006", bold=True)
GREEN_FONT = Font(color="006100", bold=True)
BOLD_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

def create_dashboard(graded_path, db_path, output_path):
    log_dash.info(f"대시보드 생성 시작: {graded_path}")

    if not os.path.exists(graded_path):
        log_dash.error(f"채점 파일을 찾을 수 없음: {graded_path}")
        print(f"오류: 채점 파일을 찾을 수 없음: {graded_path}")
        return
    if not os.path.exists(db_path):
        log_dash.error(f"DB 파일을 찾을 수 없음: {db_path}")
        print(f"오류: DB 파일을 찾을 수 없음: {db_path}")
        return

    # --- 1. 데이터 로드 ---
    try:
        summary_df = pd.read_excel(graded_path, sheet_name="summary")
        meta_df = pd.read_excel(graded_path, sheet_name="meta")
        grading_df = pd.read_excel(graded_path, sheet_name="grading")
        summary_subject_df = pd.read_excel(graded_path, sheet_name="summary_by_subject")
        
        # grading_df 이름 변경
        if '문제id' in grading_df.columns:
            grading_df = grading_df.rename(columns={'문제id': 'problem_id'})
            
        db_df = pd.read_excel(db_path)
        
        # db_df 이름 변경 (시트 0을 읽는다고 가정)
        if '문제id' in db_df.columns:
            db_df = db_df.rename(columns={'문제id': 'problem_id'})
            
    except Exception as e:
        log_dash.error(f"파일 로드 실패: {e}")
        print(f"오류: 파일 로드 실패: {e}")
        return

    # --- 2. 기본 정보 추출 ---
    summary = summary_df.iloc[0]
    meta = meta_df.iloc[0]
    student_name = summary.get("student_name", "N/A")
    exam_id = summary.get("exam_id", "N/A")
    total_score = summary.get("score", 0)
    total_q = summary.get("total", 0)
    correct_q = summary.get("correct", 0)
    user_theta = meta.get("user_theta", "N/A")

    # --- 3. '상세 성적' 시트 데이터 가공 ---
    log_dash.info("문항별 상세 데이터 가공 중...")
    
    # 3a. DB에서 필요한 컬럼만 선택
    db_id_column = 'problem_id' 
    db_cols_to_merge = []

    optional_db_cols = ['년', '월'] 
    distractor_cols = [f'선지정답률_{i}' for i in range(1, 6)]
    
    if db_id_column in db_df.columns:
        db_cols_to_merge.append(db_id_column)
    else:
         log_dash.error(f"DB 파일에 ID 컬럼(problem_id)이 없습니다.")
         print(f"오류: DB 파일에 ID 컬럼(problem_id)이 없습니다.")
         return
    
    for col in optional_db_cols + distractor_cols:
        if col in db_df.columns:
            db_cols_to_merge.append(col)
        else:
            log_dash.warning(f"DB에 '{col}' 컬럼이 없어 제외됩니다.")
            
    db_subset_df = db_df[db_cols_to_merge].copy()

    # 3b. ID 형식 통일 (merge 대비)
    grading_df['problem_id'] = grading_df['problem_id'].astype(str)
    db_subset_df['problem_id'] = db_subset_df['problem_id'].astype(str)

    # 3c. Merge
    detail_df = pd.merge(grading_df, db_subset_df, on="problem_id", how="left")
    
    # --- 선지정답률을 %로 변환 ---
    distractor_percent_cols = []
    for col in distractor_cols:
         if col in detail_df.columns:
            new_col_name = f"{col} (%)"
            # 소수점 * 100 -> 퍼센트 (1자리 반올림)
            detail_df[new_col_name] = (detail_df[col] * 100).round(1)
            detail_df.drop(columns=[col], inplace=True)
            distractor_percent_cols.append(new_col_name)
    # ---------------------------------

    # 3d. 시험 문제 번호 추가 및 컬럼 정리
    detail_df.insert(0, '시험 문제 번호', range(1, len(detail_df) + 1))
    
    final_cols_order = [
        '시험 문제 번호', 'subject', 'problem_type', 
        'answer_num', 'student_answer_num', 'is_correct'
    ]
    for col in optional_db_cols:
        if col in detail_df.columns:
            final_cols_order.append(col)

    for col in distractor_percent_cols:
         if col in detail_df.columns:
            final_cols_order.append(col)
            
    detail_df = detail_df[final_cols_order].copy()
    
    # 컬럼명 한글화
    detail_df.rename(columns={
        'subject': '지문 유형',
        'problem_type': '문제 유형',
        'answer_num': '정답',
        'student_answer_num': '제출 답안',
        'is_correct': '정답 여부'
    }, inplace=True)

    # --- 4. 'Dashboard' 시트 데이터 가공 ---
    
    # 4a. 대분류 (독서/문학) 분석
    def categorize(subject):
        if subject in LIT_SUBJECTS: return "문학"
        if subject in READ_SUBJECTS: return "독서"
        return "기타"
    
    grading_df['대분류'] = grading_df['subject'].apply(categorize)
    category_summary = grading_df.groupby('대분류')['is_correct'].agg(
        total='count',
        correct='sum'
    ).reset_index()
    category_summary['정답률(%)'] = round(category_summary['correct'] / category_summary['total'] * 100, 1)

    # 4b. 취약 유형 분석
    weak_passages = summary_subject_df[summary_subject_df['accuracy(%)'] < WEAK_PASSAGE_THRESHOLD]['subject'].tolist()
    
    problem_analysis = grading_df.groupby('problem_type')['is_correct'].agg(total='count', correct='sum')
    problem_analysis['accuracy(%)'] = (problem_analysis['correct'] / problem_analysis['total']) * 100
    weak_problems = problem_analysis[problem_analysis['accuracy(%)'] < WEAK_PROBLEM_THRESHOLD].index.tolist()

    # --- 5. Excel 파일 생성 및 'Dashboard' 시트 작성 ---
    log_dash.info("Excel 파일 생성 및 'Dashboard' 시트 작성 중...")
    
    wb = openpyxl.Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    # 5a. 헤더 정보
    ws_dash['B2'] = "시험 분석 대시보드"
    ws_dash['B2'].font = Font(size=24, bold=True)
    ws_dash['B4'] = "학생명"; ws_dash['C4'] = student_name
    ws_dash['B5'] = "시험 ID"; ws_dash['C5'] = exam_id
    ws_dash['B7'] = "총점"; ws_dash['C7'] = f"{total_score} 점"
    ws_dash['B8'] = "문항 수"; ws_dash['C8'] = f"{correct_q} / {total_q} (개)"
    ws_dash['B9'] = "학생 능력치(Theta)"; ws_dash['C9'] = user_theta
    for r in [4, 5, 7, 8, 9]: ws_dash[f'B{r}'].font = BOLD_FONT

    # 5b. 영역별 요약 (문학/독서)
    start_row = 12
    ws_dash[f'B{start_row}'] = "영역별 성취도 요약"
    ws_dash[f'B{start_row}'].font = Font(size=14, bold=True)
    headers = ["영역", "총 문항", "맞은 문항", "정답률(%)"]
    for c, header in enumerate(headers, 2):
        cell = ws_dash.cell(row=start_row+1, column=c, value=header)
        cell.font = BOLD_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER_ALIGN
    
    for r_idx, row in category_summary.iterrows():
        ws_dash.cell(row=start_row+2+r_idx, column=2, value=row['대분류'])
        ws_dash.cell(row=start_row+2+r_idx, column=3, value=row['total'])
        ws_dash.cell(row=start_row+2+r_idx, column=4, value=row['correct'])
        ws_dash.cell(row=start_row+2+r_idx, column=5, value=row['정답률(%)'])

    # 5c. 지문 유형별 정답률
    start_row = 12
    ws_dash[f'G{start_row}'] = "지문 유형별 정답률"
    ws_dash[f'G{start_row}'].font = Font(size=14, bold=True)
    headers = ["지문 유형", "총 문항", "맞은 문항", "정답률(%)"]
    for c, header in enumerate(headers, 7):
        cell = ws_dash.cell(row=start_row+1, column=c, value=header)
        cell.font = BOLD_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER_ALIGN

    for r_idx, row in summary_subject_df.iterrows():
        ws_dash.cell(row=start_row+2+r_idx, column=7, value=row['subject'])
        ws_dash.cell(row=start_row+2+r_idx, column=8, value=row['total'])
        ws_dash.cell(row=start_row+2+r_idx, column=9, value=row['correct'])
        ws_dash.cell(row=start_row+2+r_idx, column=10, value=row['accuracy(%)'])
        
    # 5d. 지문 유형별 차트
    chart = BarChart()
    chart.title = "지문 유형별 정답률 (%)"
    data_range = Reference(ws_dash, min_col=10, min_row=start_row+2, max_row=start_row+1+len(summary_subject_df))
    cat_range = Reference(ws_dash, min_col=7, min_row=start_row+2, max_row=start_row+1+len(summary_subject_df))
    chart.add_data(data_range, titles_from_data=False)
    chart.set_categories(cat_range)
    chart.legend = None
    chart.y_axis.title = "정답률(%)"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    
    ws_dash.add_chart(chart, f"L{start_row}")

    # 5e. 취약 유형
    start_row = max(start_row + len(category_summary) + 4, start_row + len(summary_subject_df) + 4)
    ws_dash[f'B{start_row}'] = "주요 취약점"
    ws_dash[f'B{start_row}'].font = Font(size=14, bold=True)
    
    ws_dash[f'B{start_row+1}'] = "취약 지문 유형 (정답률 70% 미만)"
    ws_dash[f'B{start_row+1}'].font = BOLD_FONT
    if weak_passages:
        for i, p in enumerate(weak_passages, 1):
            ws_dash.cell(row=start_row+1+i, column=2, value=p)
    else:
        ws_dash.cell(row=start_row+2, column=2, value="없음")

    ws_dash[f'G{start_row+1}'] = "취약 문제 유형 (정답률 60% 미만)"
    ws_dash[f'G{start_row+1}'].font = BOLD_FONT
    if weak_problems:
        for i, p in enumerate(weak_problems, 1):
            ws_dash.cell(row=start_row+1+i, column=7, value=p)
    else:
        ws_dash.cell(row=start_row+2, column=7, value="없음")
        
    # 컬럼 너비 조정
    for col in ['B', 'C', 'G', 'H', 'I', 'J']:
        ws_dash.column_dimensions[col].width = 15

    # --- 6. '상세 성적' 시트 작성 및 스타일링 ---
    log_dash.info("'상세 성적' 시트 작성 및 하이라이팅 적용 중...")
    
    ws_detail = wb.create_sheet(title="상세 성적")
    
    # 6a. DataFrame to Excel
    from openpyxl.utils.dataframe import dataframe_to_rows
    for r_idx, row in enumerate(dataframe_to_rows(detail_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_detail.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1: # Header
                cell.font = BOLD_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER_ALIGN
            if isinstance(value, (int, float)):
                cell.alignment = CENTER_ALIGN

    # 6b. 컬럼 인덱스 매핑 (스타일링 편의용)
    col_map = {cell.value: cell.column for cell in ws_detail[1]}
    
    # 6c. 하이라이팅 적용
    distractor_col_indices = {} # (예: 1 -> 'K')
    
    # '% (%)'가 붙은 새 컬럼명으로 매핑
    for i in range(1, 6):
        col_name = f'선지정답률_{i} (%)'
        if col_name in col_map:
            distractor_col_indices[str(i)] = col_map[col_name]
            
    if not distractor_col_indices:
        log_dash.warning("선지정답률 컬럼이 없어 하이라이팅을 스킵합니다.")
    
    else:
        for row in ws_detail.iter_rows(min_row=2, max_row=ws_detail.max_row):
            row_cells = {col: row[idx - 1] for col, idx in col_map.items()}
            
            try:
                is_correct = row_cells['정답 여부'].value
                correct_ans = str(row_cells['정답'].value)
                student_ans = str(row_cells['제출 답안'].value)

                # 1. 정답 선지 (초록색 배경)
                if correct_ans in distractor_col_indices:
                    col_idx = distractor_col_indices[correct_ans]
                    row[col_idx - 1].fill = GREEN_FILL
                
                # 2. 학생 선택
                if student_ans in distractor_col_indices:
                    col_idx = distractor_col_indices[student_ans]
                    
                    if is_correct == 1:
                        row[col_idx - 1].font = GREEN_FONT
                    else:
                        row[col_idx - 1].font = RED_FONT

                # 3. 틀린 문제 행 (빨간색 배경)
                if is_correct == 0:
                    for cell in row:
                        cell.fill = RED_FILL
                        
            except Exception as e:
                log_dash.warning(f"{row[0].row}행 스타일링 중 오류: {e}")

    # 6d. '상세 성적' 시트 너비/고정 조정
    for i, col in enumerate(detail_df.columns, 1):
        try:
            width = max(len(str(col)), detail_df[col].astype(str).map(len).max()) * 1.1
            ws_detail.column_dimensions[get_column_letter(i)].width = min(max(width, 12), 30)
        except Exception:
            ws_detail.column_dimensions[get_column_letter(i)].width = 15
    ws_detail.freeze_panes = 'A2' # 헤더 고정

    # --- 7. 저장 ---
    try:
        wb.save(output_path)
        log_dash.info(f"대시보드 생성 완료: {os.path.abspath(output_path)}")
        print(f"대시보드 생성 완료: {os.path.abspath(output_path)}")
    except PermissionError:
        log_dash.error(f"권한 오류: 파일을 저장할 수 없습니다. 파일이 열려있는지 확인하세요. ({output_path})")
        print(f"권한 오류: 파일을 저장할 수 없습니다. 파일이 열려있는지 확인하세요. ({output_path})")
    except Exception as e:
        log_dash.error(f"파일 저장 실패: {e}")
        print(f"파일 저장 실패: {e}")


if __name__ == "__main__":
    # (주의) 실행 전 상단의 '설정' 3가지를 꼭 확인하세요!
    create_dashboard(
        graded_path=GRADED_FILE_PATH,
        db_path=DB_PATH,
        output_path=DASHBOARD_OUTPUT_PATH
    )
